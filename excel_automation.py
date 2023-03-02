import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import re
import os


""" creating a dictionary to replace characters
key - character to be replaced
value - character to replace with """

to_remov = {"period": ".", "dot": ".", "full stop": ".", "question mark": "?", "exclamation point": "!", "comma": ",", "colon": ":", "semicolon": ";", 
"dash": "-", "hyphen": "-", "left bracket": "[",  "right bracket": "]", "left brace": "{", "right brace": "}", 
"left parenthese": "(", "righr parenthese": ")", "backslash": "\\", "slash": "/", "underscore": "_", "plus": "+", 
"single quotation mark": "'", "double quotation mark": '"', "ellipsis": "..."}

# extract key information from speech based on the features of the spreadsheet 
def extract_features(all_features, string):
    for char in to_remov.keys():
        string = string.replace(char, to_remov[char])
    val_dict = dict.fromkeys(all_features)
    val_dict['add feature'] = []
    idxs = []
    for h in all_features:    
        idxs += [i.start() for i in re.finditer(h, string)]
    idxs.sort()

    for i in range(len(idxs)):
        from_idx = idxs[i]
        to_idx = idxs[i+1] if i+1 != len(idxs) else len(string)
        sub_str = string[from_idx:to_idx].strip()
        feature = ''
        for f in all_features:
            if f in sub_str: 
                feature = f
                break 
        clean_sub_str = sub_str.split(feature)[1].strip()
        if clean_sub_str:
            if 'id' in feature or 'barcode' in feature:
                clean_sub_str = clean_sub_str.replace(" ","")
            if feature != 'add feature':
                val_dict[feature] = clean_sub_str
            else:
                val_dict[feature].append(clean_sub_str)
    
    return val_dict

# check if the new input is whether supplementary information (if it is, add to the existed record; if it is not, add to a new record)
def check_additionalinfo(sheet, val_dict):
    addition = False
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
        for i in range(len(list(val_dict.keys()))-1):                
            if 'id' in list(val_dict.keys())[i]:
                if ((list(val_dict.values())[i] is not None) and (row[i].value is not None) and (list(val_dict.values())[i].upper() == row[i].value)):
                    addition = True
                else:
                    addition = False
                    break 
    if addition:
        return row_idx+2  
    else:
        return sheet.max_row+1

# import text in transcript to spreadsheet
def script_to_spreadsheet(header='', transcript_name='', spreadsheet_name='', sheet_name=''):

    # header = ['Cell_ID', 'Plate_barcode', 'Plate_ID', 'Well_barcode', 'Well_ID', 'Comment']

    # check if create a new spreadsheet or overwrite an existed one
    if not os.path.isfile(f"{spreadsheet_name}.xlsx"):
        wb = xl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        header = header.split(',')
        for i in range(len(header)):
            cell = sheet.cell(1, i+1)
            cell.value = header[i]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        wb = xl.load_workbook(f'{spreadsheet_name}.xlsx')
        sheet = wb[sheet_name]
        header = [cell.value for cell in sheet[1]] 
        
    with open(f'{transcript_name}.txt') as f:
        lines = f.readlines()
    for line in lines:
        all_features = [i.lower().replace('_', ' ') for i in header] + ['add feature'] # add extra features for different analysis purposes
        val_dict = extract_features(all_features, line.strip().lower())
        while val_dict['add feature']:
            if val_dict['add feature'][0] and val_dict['add feature'][0] not in header:
                header += [val_dict['add feature'][0]]
                cell = sheet.cell(1, len(header))
                cell.value = header[len(header)-1]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            val_dict['add feature'].pop(0)
        if all(value is None for value in list(val_dict.values())[:-1]):
            continue
        row_idx = check_additionalinfo(sheet, val_dict)
        for i in range(len(val_dict)-1):
            cell = sheet.cell(row_idx, i+1)
            if list(val_dict.values())[i] is None:
                if cell.value is None:
                    cell.value = None
            elif list(val_dict.keys())[i] == 'comment':
                if cell.value is None:
                    cell.value = list(val_dict.values())[i].capitalize() + '.' if (not list(val_dict.values())[i].endswith('.')) else list(val_dict.values())[i].capitalize()
                else:
                    cell.value += ' ' + list(val_dict.values())[i].capitalize() + '.' if (not list(val_dict.values())[i].endswith('.')) else list(val_dict.values())[i].capitalize()
            else:
                if 'id' in list(val_dict.keys())[i] or 'barcode' in list(val_dict.keys())[i]:
                    cell.value = list(val_dict.values())[i].upper()
                else:
                    cell.value = list(val_dict.values())[i]
    
    wb.save(filename=f'{spreadsheet_name}.xlsx')

# def open_spreadsheet(spreadsheet_name):
#     wb = xl.load_workbook(f'{spreadsheet_name}.xlsx')
#     return wb

# edit the cell in spreadsheet
def edit_spreadsheet_cell(spreadsheet_name, sheet_name, bylocation, cell_loc='', new_value=''):

    # Load the spreadsheet file
    if not os.path.isfile(f"{spreadsheet_name}.xlsx"):
        wb = xl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
    else:
        wb = xl.load_workbook(f'{spreadsheet_name}.xlsx')
        sheet = wb[sheet_name]    

    editable = True
    if bylocation:
        # Edit the specified cell
        sheet[f'{cell_loc}'].value = new_value
        #sheet.cell(row=cell_row, column=cell_column).value = new_value
    else:
        header  = [cell.column for cell in sheet[1]]
        all_features = [i.lower().replace('_', ' ') for i in header]
        val_dict = extract_features(all_features, cell_loc.strip().lower())  
        row_idx = check_additionalinfo(sheet, val_dict)
        if row_idx != sheet.max_row:
            col_idx = all_features.index(list(val_dict.keys())[list(val_dict.values()).index('')])
            cell = sheet.cell(row_idx, col_idx)
            cell.value = new_value
        else:
            editable = False
    
    # Save the changes
    wb.save(filename=f'{spreadsheet_name}.xlsx')
    return editable